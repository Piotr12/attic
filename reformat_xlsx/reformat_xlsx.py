"""
helper function for parsing an xlsx that has plenty of merged cells added to look 'nice'
"""
from openpyxl import load_workbook

def reformat(file, separator, none_list=[None]):
    """
    A | B | C | D
    1 | 2 | 3 | 4
      | a | 3 | 4
      |   | 3 | 4
    Will get an 1 added in second row and "1 | a" in third.
    What is considered empty is actually when cell.value==None unless you provide an optional list of what is considered empty (e.g. (None,""))
    """
    output = ""
    workbook = load_workbook(file)
    worksheet = workbook.active
    last_row = []
    for row in worksheet.iter_rows():
        # either a header row ... and I assume its full of valuable data
        if len(last_row) == 0:
            for cell in row:
                last_row.append(cell.value)
                output += cell.value + separator
            output = output[:-1] + "\n"
        # or a data one, where I will add values from previous row if None found
        else:
            column = 0
            line = ""
            for cell in row:
                if not cell.value in none_list:
                    last_row[column] = cell.value
                line += str(last_row[column]) + separator
                column += 1
            output += line[:-1] + "\n"
    return output
